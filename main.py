from bs4 import BeautifulSoup
import requests
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font
import re

base_url = 'https://cv.lv/lv/search?limit=20&offset={}&categories%5B0%5D=INFORMATION_TECHNOLOGY&fuzzy=true&suitableForRefugees=false&isHourlySalary=false&isRemoteWork=false&isQuickApply=false'
offset = 0

# Create three empty DataFrames to store job information
df1 = pd.DataFrame(columns=['Job Title', 'Employer', 'Salary', 'Deadline', 'Link']) # For junior vacancies
df2 = pd.DataFrame(columns=['Job Title', 'Employer', 'Salary', 'Deadline', 'Link']) # For salary range vacancies
df3 = pd.DataFrame(columns=['Job Title', 'Employer', 'Salary', 'Deadline', 'Link']) # For hourly rate vacancies

while True:
    html_text = requests.get(base_url.format(offset)).text
    soup = BeautifulSoup(html_text, 'lxml')
    jobs = soup.find_all('div', class_='jsx-3024910437')

    if not jobs:
        break
    
    for job in jobs:
        title = job.find('span', class_='jsx-3024910437 vacancy-item__title')
        link = job.find('a', class_='jsx-3024910437 vacancy-item')
        salary = job.find('span', class_='jsx-3024910437 vacancy-item__info-labels')
        employer = job.find('div', class_='jsx-3024910437 vacancy-item__column')
        deadline = job.find('span', class_='jsx-3024910437 vacancy-item__expiry')
        
        if title and link:
            job_title = title.text
            job_link = 'https://cv.lv' + link['href']
            job_salary = salary.text
            job_employer = employer.text
            job_deadline = (deadline.text).split(":")[-1].strip()
            if 'junior' in job_title.lower() or 'jaunākais' in job_title.lower():
                df1 = df1._append({'Job Title': job_title, 'Employer': job_employer, 'Salary': job_salary, 'Deadline': job_deadline, 'Link': job_link}, ignore_index=True)
            if 'junior' not in job_title.lower() or 'jaunākais' not in job_title.lower():
                # Parse the salary to get the start and end of the range
                matches = re.findall(r'(\d+)', job_salary)
                if matches:
                    start_salary = int(matches[0])
                    if 400 <= start_salary <= 1200:
                        if len(matches) > 1:
                            end_salary = int(matches[1])
                            if 700 <= end_salary <= 1600:
                                df2 = df2._append({'Job Title': job_title, 'Employer': job_employer, 'Salary': job_salary, 'Deadline': job_deadline, 'Link': job_link}, ignore_index=True)
            # If the salary is given per hour
            if '/st.' in job_salary and ('junior' not in job_title.lower() or 'jaunākais' not in job_title.lower()):
                df3 = df3._append({'Job Title': job_title, 'Employer': job_employer, 'Salary': job_salary, 'Deadline': job_deadline, 'Link': job_link}, ignore_index=True)

    # Increment the offset by 20 (as each page displays 20 jobs)
    offset += 20
    
wb = Workbook()
del wb['Sheet']

ws1 = wb.create_sheet('Junior vacancies')
ws2 = wb.create_sheet('Salary range vacancies')
ws3 = wb.create_sheet('Hourly rate vacancies')

for row in dataframe_to_rows(df1, index=False, header=True):
    ws1.append(row)    
for row in dataframe_to_rows(df2, index=False, header=True):
    ws2.append(row) 
for row in dataframe_to_rows(df3, index=False, header=True):
    ws3.append(row)

font = Font(bold=True)

for ws in [ws1, ws2, ws3]:
    for column_cells in ws.columns:
        # Adjust the width of the column based on the length of the longest cell
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length
        if column_cells[0].row == 1:
            column_cells[0].font = font

wb.save('jobs_scraping/jobs/scraped_jobs.xlsx')  
